"""Generate production lists as a multi-sheet Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AMBER_HEX = "D97706"
DEEP_FORGE_HEX = "36363E"
HEADER_FILL = PatternFill(start_color=DEEP_FORGE_HEX, end_color=DEEP_FORGE_HEX, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=9)
ALT_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

MEMBER_LABELS = {
    "frame_top": "Bovendorpel", "frame_bottom": "Onderdorpel",
    "frame_left": "Stijl links", "frame_right": "Stijl rechts",
    "divider_h": "Tussendorpel", "divider_v": "Tussenstijl",
    "sash_top": "Raamhout boven", "sash_bottom": "Raamhout onder",
    "sash_left": "Raamhout links", "sash_right": "Raamhout rechts",
}

GASKET_LABELS = {
    "glazing_inner": "Binnenrubber beglazing", "glazing_outer": "Buitenrubber beglazing",
    "sash_seal": "Vleugelafdichting", "frame_seal": "Kozijnafdichting",
}


def generate_production_xlsx(production_data_list: list, output_path: str):
    """Generate an Excel workbook with production lists."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Flatten all kozijnen into combined lists
    all_cut = []
    all_glass = []
    all_hw = []
    all_gasket = []
    all_panel = []
    all_bom = []

    for prod in production_data_list:
        mark = prod.get("kozijnMark", "?")
        for item in prod.get("cutList", []):
            item["_mark"] = mark
            all_cut.append(item)
        for item in prod.get("glassList", []):
            item["_mark"] = mark
            all_glass.append(item)
        for item in prod.get("hardwareList", []):
            item["_mark"] = mark
            all_hw.append(item)
        for item in prod.get("gasketList", []):
            item["_mark"] = mark
            all_gasket.append(item)
        for item in prod.get("panelList", []):
            item["_mark"] = mark
            all_panel.append(item)
        for item in prod.get("bom", []):
            item["_mark"] = mark
            all_bom.append(item)

    # Kortlijst
    ws = wb.create_sheet("Kortlijst")
    headers = ["Kozijn", "Pos.", "Onderdeel", "Profiel", "Materiaal",
               "Netto (mm)", "Bruto (mm)", "Hoek L", "Hoek R", "Aantal"]
    _write_headers(ws, headers)
    for i, item in enumerate(all_cut):
        member = MEMBER_LABELS.get(item.get("memberType", ""), item.get("memberType", ""))
        row = [item["_mark"], item.get("pieceId", ""), member,
               item.get("profileName", ""), item.get("material", ""),
               round(item.get("netLengthMm", 0)), round(item.get("grossLengthMm", 0)),
               f'{item.get("miterLeftDeg", 90):.0f}\u00b0',
               f'{item.get("miterRightDeg", 90):.0f}\u00b0',
               item.get("quantity", 1)]
        _write_row(ws, i + 2, row)
    _autofit(ws)

    # Glaslijst
    ws = wb.create_sheet("Glaslijst")
    headers = ["Kozijn", "Pos.", "Glastype", "Breedte", "Hoogte", "Dikte", "Ug", "Opp. (m\u00b2)", "Aantal"]
    _write_headers(ws, headers)
    for i, item in enumerate(all_glass):
        row = [item["_mark"], item.get("pieceId", ""), item.get("glassType", ""),
               round(item.get("widthMm", 0)), round(item.get("heightMm", 0)),
               round(item.get("thicknessMm", 0)), round(item.get("ugValue", 0), 1),
               round(item.get("areaM2", 0), 2), item.get("quantity", 1)]
        _write_row(ws, i + 2, row)
    _autofit(ws)

    # Beslaglijst
    ws = wb.create_sheet("Beslaglijst")
    headers = ["Kozijn", "Cel", "Component", "Omschrijving", "Aantal"]
    _write_headers(ws, headers)
    for i, item in enumerate(all_hw):
        row = [item["_mark"], item.get("cellIndex", 0) + 1,
               item.get("component", ""), item.get("description", ""),
               item.get("quantity", 1)]
        _write_row(ws, i + 2, row)
    _autofit(ws)

    # Rubberlijst
    ws = wb.create_sheet("Rubberlijst")
    headers = ["Kozijn", "Type", "Lengte (mm)", "Aantal"]
    _write_headers(ws, headers)
    for i, item in enumerate(all_gasket):
        label = GASKET_LABELS.get(item.get("gasketType", ""), item.get("gasketType", ""))
        row = [item["_mark"], label, round(item.get("lengthMm", 0)), item.get("quantity", 1)]
        _write_row(ws, i + 2, row)
    _autofit(ws)

    # Paneellijst
    if all_panel:
        ws = wb.create_sheet("Paneellijst")
        headers = ["Kozijn", "Pos.", "Breedte", "Hoogte", "Type", "Aantal"]
        _write_headers(ws, headers)
        for i, item in enumerate(all_panel):
            row = [item["_mark"], item.get("pieceId", ""),
                   round(item.get("widthMm", 0)), round(item.get("heightMm", 0)),
                   item.get("panelType", ""), item.get("quantity", 1)]
            _write_row(ws, i + 2, row)
        _autofit(ws)

    # Stuklijst
    ws = wb.create_sheet("Stuklijst")
    headers = ["Kozijn", "Categorie", "Omschrijving", "Eenheid", "Hoeveelheid"]
    _write_headers(ws, headers)
    for i, item in enumerate(all_bom):
        row = [item["_mark"], item.get("category", ""), item.get("description", ""),
               item.get("unit", ""), round(item.get("quantity", 0), 2)]
        _write_row(ws, i + 2, row)
    _autofit(ws)

    wb.save(output_path)


def _write_headers(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _write_row(ws, row_num, values):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if row_num % 2 == 0:
            cell.fill = ALT_FILL


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)
